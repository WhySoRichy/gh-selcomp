"""
Extractor de datos de Hojas de Vida usando Groq AI (Llama 3.1)
Portal Gestión Humana - Selcomp

Modelo: llama-3.1-8b-instant
Límites: 14,400 requests/día, 30 requests/min
"""

import os
import json
import time
import fitz  # PyMuPDF para extraer texto de PDFs
from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# OCR para PDFs basados en imagen (Canva, diseños gráficos, etc.)
try:
    import pytesseract
    from PIL import Image
    import io
    OCR_DISPONIBLE = True
except ImportError:
    OCR_DISPONIBLE = False

# ============================================
# CONFIGURACIÓN
# ============================================
# Leer API Key desde variable de entorno o archivo .env
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

# Intentar cargar desde .env si no está en el entorno
if not GROQ_API_KEY:
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                if line.startswith("GROQ_API_KEY="):
                    GROQ_API_KEY = line.split("=", 1)[1].strip().strip('"\'')
                    break

if not GROQ_API_KEY:
    raise ValueError("GROQ_API_KEY no configurada. Agrégala a las variables de entorno o al archivo .env")

MODELO = "llama-3.1-8b-instant"  # 14,400 requests/día, 30 RPM

# Ruta del Excel de Prospectos
RUTA_PROSPECTOS = os.path.join(os.path.dirname(__file__), "Prospectos.xlsx")

# Configurar Tesseract OCR
TESSERACT_PATH = os.getenv("TESSERACT_PATH", "")
if not TESSERACT_PATH:
    # Intentar cargar desde .env
    env_path_tess = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path_tess):
        with open(env_path_tess, "r") as f:
            for line in f:
                if line.startswith("TESSERACT_PATH="):
                    TESSERACT_PATH = line.split("=", 1)[1].strip().strip("\"'")
                    break

# Rutas comunes de Tesseract en Windows
if not TESSERACT_PATH:
    rutas_comunes = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for ruta in rutas_comunes:
        if os.path.exists(ruta):
            TESSERACT_PATH = ruta
            break

if OCR_DISPONIBLE and TESSERACT_PATH:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

# Umbral mínimo de caracteres para considerar extracción válida
UMBRAL_TEXTO_MINIMO = 50

# ============================================
# FUNCIONES PRINCIPALES
# ============================================

def configurar_groq():
    """Configura la conexión con Groq AI"""
    return Groq(api_key=GROQ_API_KEY)


def extraer_texto_pdf(ruta_pdf):
    """Extrae todo el texto de un PDF.
    
    Estrategia:
    1. Intenta extracción directa de texto con PyMuPDF
    2. Si el texto es insuficiente (<50 chars), usa OCR con Tesseract
       (ideal para PDFs de Canva, diseños gráficos, imágenes escaneadas)
    """
    try:
        doc = fitz.open(ruta_pdf)
        
        # Paso 1: Extracción directa de texto
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()
        texto = texto.strip()
        
        # Si el texto es suficiente, retornarlo
        if len(texto) >= UMBRAL_TEXTO_MINIMO:
            doc.close()
            return texto
        
        # Paso 2: Fallback OCR para PDFs basados en imagen
        print(f"   ⚠️ Texto directo insuficiente ({len(texto)} chars). Intentando OCR...")
        
        if not OCR_DISPONIBLE:
            print(f"   ❌ OCR no disponible. Instalar: pip install pytesseract Pillow")
            doc.close()
            return texto if texto else None
        
        if not TESSERACT_PATH or not os.path.exists(TESSERACT_PATH):
            print(f"   ❌ Tesseract no encontrado. Instalar desde: https://github.com/UB-Mannheim/tesseract")
            doc.close()
            return texto if texto else None
        
        texto_ocr = ""
        for num_pagina, pagina in enumerate(doc):
            try:
                # Renderizar página como imagen a 300 DPI para mejor OCR
                mat = fitz.Matrix(300/72, 300/72)  # 300 DPI
                pix = pagina.get_pixmap(matrix=mat)
                img_bytes = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_bytes))
                
                # OCR con Tesseract (español + inglés)
                texto_pagina = pytesseract.image_to_string(img, lang='spa+eng')
                texto_ocr += texto_pagina + "\n"
            except Exception as e:
                print(f"   ⚠️ Error OCR en página {num_pagina + 1}: {e}")
                continue
        
        doc.close()
        texto_ocr = texto_ocr.strip()
        
        if texto_ocr:
            print(f"   ✅ OCR exitoso: {len(texto_ocr)} caracteres extraídos")
            return texto_ocr
        
        # Si OCR tampoco funcionó, devolver lo que tengamos
        return texto if texto else None
        
    except Exception as e:
        print(f"Error al leer PDF {ruta_pdf}: {e}")
        return None


def analizar_hv_con_groq(client, texto_hv, max_reintentos=3):
    """Envía el texto de la HV a Groq y extrae los datos"""
    
    prompt = f"""Extrae estos datos de la hoja de vida. Responde SOLO JSON válido, sin explicaciones ni texto adicional.

Datos a extraer:
- nombres (solo nombres, sin apellidos)
- apellidos (solo apellidos)
- educacion (título profesional o técnico más reciente)
- anos_experiencia (número entero de años de experiencia total)
- experiencia_laboral (resumen breve de cargos y empresas, máximo 100 caracteres)

Si no encuentras un dato, pon null. Para anos_experiencia pon 0 si no hay experiencia.

HOJA DE VIDA:
{texto_hv}

JSON:"""

    for intento in range(max_reintentos):
        try:
            respuesta = client.chat.completions.create(
                model=MODELO,
                messages=[
                    {"role": "system", "content": "Eres un extractor de datos de hojas de vida. Responde SOLO con JSON válido."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=500
            )
            texto_respuesta = respuesta.choices[0].message.content.strip()
            
            # Limpiar si viene con markdown
            if texto_respuesta.startswith("```"):
                texto_respuesta = texto_respuesta.split("```")[1]
                if texto_respuesta.startswith("json"):
                    texto_respuesta = texto_respuesta[4:]
                texto_respuesta = texto_respuesta.strip()
            
            datos = json.loads(texto_respuesta)
            return datos
        
        except json.JSONDecodeError as e:
            print(f"Error al parsear JSON: {e}")
            print(f"Respuesta recibida: {texto_respuesta[:500]}")
            return None
        except Exception as e:
            error_str = str(e)
            # Si es error de rate limit, esperar y reintentar
            if "rate_limit" in error_str.lower() or "429" in error_str:
                if intento < max_reintentos - 1:
                    tiempo_espera = (intento + 1) * 10  # 10s, 20s, 30s
                    print(f"Rate limit. Esperando {tiempo_espera}s ({intento + 1}/{max_reintentos})...")
                    time.sleep(tiempo_espera)
                    continue
            print(f"Error al comunicarse con Groq: {e}")
            return None
    
    return None


def crear_excel(datos_hvs, ruta_salida):
    """Agrega los datos extraídos al Excel Prospectos existente"""
    from openpyxl import load_workbook
    
    # Cargar Excel existente o crear uno nuevo
    if os.path.exists(RUTA_PROSPECTOS):
        wb = load_workbook(RUTA_PROSPECTOS)
        ws = wb.active
        print(f"📂 Cargando Excel existente: {RUTA_PROSPECTOS}")
        # Encontrar la última fila con datos
        ultima_fila = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Prospectos"
        # Crear headers (5 columnas)
        headers = ["Nombres", "Apellidos", "Educacion", "AñosdeExperiencia", "Experiencia Laboral"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        ultima_fila = 2
        print(f"📄 Creando nuevo Excel: {RUTA_PROSPECTOS}")
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir datos (5 columnas)
    for datos in datos_hvs:
        ws.cell(row=ultima_fila, column=1, value=datos.get('nombres', '')).border = thin_border
        ws.cell(row=ultima_fila, column=2, value=datos.get('apellidos', '')).border = thin_border
        ws.cell(row=ultima_fila, column=3, value=datos.get('educacion', '')).border = thin_border
        ws.cell(row=ultima_fila, column=4, value=datos.get('anos_experiencia', 0)).border = thin_border
        ws.cell(row=ultima_fila, column=5, value=datos.get('experiencia_laboral', '')).border = thin_border
        ultima_fila += 1
    
    # Guardar
    wb.save(RUTA_PROSPECTOS)
    print(f"✅ Excel actualizado: {RUTA_PROSPECTOS}")


def procesar_carpeta_hvs(carpeta_pdfs, ruta_excel_salida):
    """Procesa todos los PDFs de una carpeta"""
    
    print("=" * 50)
    print("🚀 EXTRACTOR DE HOJAS DE VIDA - Groq AI")
    print("=" * 50)
    
    # Buscar PDFs
    archivos_pdf = [f for f in os.listdir(carpeta_pdfs) if f.lower().endswith('.pdf')]
    
    if not archivos_pdf:
        print(f"❌ No se encontraron PDFs en: {carpeta_pdfs}")
        return
    
    print(f"📁 Encontrados {len(archivos_pdf)} PDFs para procesar")
    
    # Configurar Groq
    print("🔌 Conectando con Groq AI...")
    cliente = configurar_groq()
    
    # Procesar cada PDF
    datos_extraidos = []
    
    for i, archivo in enumerate(archivos_pdf, 1):
        ruta_completa = os.path.join(carpeta_pdfs, archivo)
        print(f"\n[{i}/{len(archivos_pdf)}] Procesando: {archivo}")
        
        # Extraer texto
        texto = extraer_texto_pdf(ruta_completa)
        if not texto:
            print(f"   ⚠️ No se pudo extraer texto")
            continue
        
        print(f"   📄 Texto extraído: {len(texto)} caracteres")
        
        # Analizar con Groq
        datos = analizar_hv_con_groq(cliente, texto)
        if datos:
            datos["_archivo_origen"] = archivo
            datos_extraidos.append(datos)
            nombre_mostrar = f"{datos.get('nombres', '')} {datos.get('apellidos', '')}".strip() or "Sin nombre"
            print(f"   ✅ Datos extraídos: {nombre_mostrar}")
        else:
            print(f"   ❌ Error al analizar")
        
        # Esperar entre requests para evitar rate limit
        if i < len(archivos_pdf):
            time.sleep(5)  # 5 segundos entre cada HV
    
    # Crear Excel
    if datos_extraidos:
        print(f"\n📊 Generando Excel con {len(datos_extraidos)} registros...")
        crear_excel(datos_extraidos, ruta_excel_salida)
    else:
        print("\n❌ No se pudieron extraer datos de ningún PDF")
    
    print("\n" + "=" * 50)
    print("✨ Proceso completado")
    print("=" * 50)


def procesar_un_pdf(ruta_pdf):
    """Procesa un solo PDF y muestra los resultados (para pruebas)"""
    
    print("=" * 50)
    print("🧪 PRUEBA - Extractor de HV")
    print("=" * 50)
    
    # Extraer texto
    print(f"\n📄 Leyendo: {ruta_pdf}")
    texto = extraer_texto_pdf(ruta_pdf)
    
    if not texto:
        print("❌ No se pudo extraer texto del PDF")
        return None
    
    print(f"✅ Texto extraído: {len(texto)} caracteres")
    print("\n--- Primeros 500 caracteres ---")
    print(texto[:500])
    print("--- Fin preview ---\n")
    
    # Analizar con Groq
    print("🤖 Enviando a Groq AI...")
    cliente = configurar_groq()
    datos = analizar_hv_con_groq(cliente, texto)
    
    if datos:
        print("\n✅ DATOS EXTRAÍDOS:")
        print("-" * 40)
        for campo, valor in datos.items():
            print(f"  {campo}: {valor}")
        return datos
    else:
        print("❌ Error al extraer datos")
        return None


def procesar_pdf_automatico(ruta_pdf, vacante_titulo=""):
    """
    Procesa un solo PDF automáticamente y lo agrega al Excel.
    Diseñado para ser llamado desde PHP después de cada postulación.
    
    Args:
        ruta_pdf: Ruta completa al archivo PDF
        vacante_titulo: Título de la vacante (opcional, para la columna Vacantes Aplicables)
    
    Returns:
        dict con resultado: {'success': bool, 'message': str, 'data': dict}
    """
    from openpyxl import load_workbook
    import logging
    
    # Configurar logging para archivo
    log_file = os.path.join(os.path.dirname(__file__), "extractor.log")
    logging.basicConfig(
        filename=log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    resultado = {'success': False, 'message': '', 'data': None}
    
    try:
        logging.info(f"Procesando PDF: {ruta_pdf}")
        
        # Verificar que el archivo existe
        if not os.path.exists(ruta_pdf):
            resultado['message'] = f"Archivo no encontrado: {ruta_pdf}"
            logging.error(resultado['message'])
            return resultado
        
        # Extraer texto del PDF
        texto = extraer_texto_pdf(ruta_pdf)
        if not texto:
            resultado['message'] = "No se pudo extraer texto del PDF"
            logging.error(resultado['message'])
            return resultado
        
        logging.info(f"Texto extraído: {len(texto)} caracteres")
        
        # Analizar con Groq
        cliente = configurar_groq()
        datos = analizar_hv_con_groq(cliente, texto)
        
        if not datos:
            resultado['message'] = "No se pudieron extraer datos con Groq"
            logging.error(resultado['message'])
            return resultado
        
        logging.info(f"Datos extraídos: {datos.get('nombres', '')} {datos.get('apellidos', '')}")
        
        # Agregar al Excel
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if os.path.exists(RUTA_PROSPECTOS):
            wb = load_workbook(RUTA_PROSPECTOS)
            ws = wb.active
            ultima_fila = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Prospectos"
            headers = ["Nombres", "Apellidos", "Educacion", "AñosdeExperiencia", "Experiencia Laboral"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            ultima_fila = 2
        
        # Escribir datos (5 columnas: Nombres, Apellidos, Educacion, AñosdeExperiencia, Experiencia Laboral)
        ws.cell(row=ultima_fila, column=1, value=datos.get('nombres', '')).border = thin_border
        ws.cell(row=ultima_fila, column=2, value=datos.get('apellidos', '')).border = thin_border
        ws.cell(row=ultima_fila, column=3, value=datos.get('educacion', '')).border = thin_border
        ws.cell(row=ultima_fila, column=4, value=datos.get('anos_experiencia', 0)).border = thin_border
        ws.cell(row=ultima_fila, column=5, value=datos.get('experiencia_laboral', '')).border = thin_border
        
        wb.save(RUTA_PROSPECTOS)
        
        resultado['success'] = True
        resultado['message'] = "PDF procesado y agregado al Excel exitosamente"
        resultado['data'] = datos
        logging.info(f"✅ Agregado a Excel: {datos.get('nombres', '')} {datos.get('apellidos', '')}")
        
    except Exception as e:
        resultado['message'] = f"Error: {str(e)}"
        logging.error(f"Error procesando PDF: {str(e)}")
    
    return resultado


# ============================================
# EJECUCIÓN
# ============================================

if __name__ == "__main__":
    import sys
    
    # Carpeta por defecto de hojas de vida
    CARPETA_HVS_DEFAULT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Documentos", "HojasDeVida")
    
    # Modo automático: python extractor_hv.py --auto "ruta_pdf" "vacante_titulo"
    if len(sys.argv) >= 3 and sys.argv[1] == "--auto":
        ruta_pdf = sys.argv[2]
        vacante_titulo = sys.argv[3] if len(sys.argv) > 3 else ""
        resultado = procesar_pdf_automatico(ruta_pdf, vacante_titulo)
        # Salida JSON para PHP
        print(json.dumps(resultado, ensure_ascii=False))
        sys.exit(0 if resultado['success'] else 1)
    
    elif len(sys.argv) > 1:
        # Si se pasa un argumento, procesar ese PDF o carpeta
        ruta = sys.argv[1]
        if os.path.isfile(ruta):
            procesar_un_pdf(ruta)
        elif os.path.isdir(ruta):
            procesar_carpeta_hvs(ruta, RUTA_PROSPECTOS)
    else:
        # Sin argumentos: procesar carpeta por defecto de HojasDeVida
        print(f"📁 Usando carpeta por defecto: {CARPETA_HVS_DEFAULT}")
        if os.path.exists(CARPETA_HVS_DEFAULT):
            procesar_carpeta_hvs(CARPETA_HVS_DEFAULT, RUTA_PROSPECTOS)
        else:
            print(f"❌ No existe la carpeta: {CARPETA_HVS_DEFAULT}")
            print("\nUso:")
            print("  python extractor_hv.py                              - Procesar carpeta HojasDeVida")
            print("  python extractor_hv.py <ruta_pdf>                   - Procesar un PDF (prueba)")
            print("  python extractor_hv.py <carpeta_pdfs>               - Procesar carpeta específica")
            print("  python extractor_hv.py --auto <pdf> <vacante>       - Modo automático (para PHP)")